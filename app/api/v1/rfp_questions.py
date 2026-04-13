"""RFP Questions API — import questions from Excel/CSV (column A) and store in rfpquestions table."""
import csv
import io
import json
import logging
import uuid
from datetime import datetime, timezone

from fastapi import APIRouter, BackgroundTasks, File, Form, HTTPException, Query, UploadFile
from openpyxl import load_workbook
from pydantic import BaseModel
from sqlalchemy import func, or_, select

from app.api.deps import DbSession, CurrentUser
from app.models.rfp_question import RFPQuestion, generate_rfpid
from app.models.user import User, UserRole
from app.services.rfp_completion_email import build_qa_excel_bytes, send_rfp_completion_notification
from app.utils.conversation_id import generate_conversation_id

logger = logging.getLogger(__name__)

router = APIRouter(prefix="/rfp-questions", tags=["rfp-questions"])


def _require_rfp_owner_or_privileged(current_user: User, row: RFPQuestion) -> None:
    if current_user.role in (UserRole.admin, UserRole.manager):
        return
    if row.user_id != current_user.id:
        raise HTTPException(status_code=403, detail="Access denied")


def _parse_collab_ids(raw: str | None) -> list[str]:
    if not raw or not str(raw).strip():
        return []
    return [x.strip() for x in str(raw).split(",") if x.strip()]


def _format_collab_ids(ids: list[str]) -> str:
    seen: set[str] = set()
    out: list[str] = []
    for i in ids:
        i = (i or "").strip()
        if not i or i in seen:
            continue
        seen.add(i)
        out.append(i)
    return ",".join(out)


def _user_in_collaborators(row: RFPQuestion, user_id: str) -> bool:
    return user_id in _parse_collab_ids(getattr(row, "collaborator_user_ids", None))


def _rfp_accessible_filter(current_user: User):
    """Owner OR collaborator (comma-padded match; no SQL LIKE wildcards in user ids)."""
    uid = current_user.id
    owned = RFPQuestion.user_id == uid
    padded = func.concat(",", func.coalesce(RFPQuestion.collaborator_user_ids, ""), ",")
    is_collab = padded.like(f"%,{uid},%")
    return or_(owned, is_collab)


def _require_rfp_access(current_user: User, row: RFPQuestion) -> None:
    if current_user.role in (UserRole.admin, UserRole.manager):
        return
    if row.user_id == current_user.id:
        return
    if _user_in_collaborators(row, current_user.id):
        return
    raise HTTPException(status_code=403, detail="Access denied")


NO_CONTEXT_MESSAGE = "Sorry No articles found"
# Frontend / search layer may prefix low-context answers (keep in sync with UploadRFP.js UNANSWERED_PREFIX).
UNANSWERED_PREFIX = "Unanswered : "


def _confidence_as_array(raw: str | list | None) -> list:
    """Parse confidence from DB; always return a list (only array). Handles MySQL JSON column returning list."""
    if raw is None:
        return []
    if isinstance(raw, list):
        return list(raw)
    if not raw:
        return []
    try:
        val = json.loads(raw)
        return list(val) if isinstance(val, list) else []
    except (json.JSONDecodeError, TypeError):
        return []


def _answers_for_response(answers: list) -> list:
    """Return answers for API response; empty/missing values become NO_CONTEXT_MESSAGE."""
    return [
        (a if (a is not None and str(a).strip()) else NO_CONTEXT_MESSAGE)
        for a in answers
    ]


def _is_successful_answer(a: str | None) -> bool:
    """True when the stored answer counts as a completed generation (not error / no-context / unanswered)."""
    s = (a or "").strip()
    if not s:
        return False
    if s.startswith("[Error:"):
        return False
    if s == NO_CONTEXT_MESSAGE:
        return False
    if s.startswith(UNANSWERED_PREFIX):
        return False
    if s.startswith("Unanswered:"):
        return False
    return True


def _all_nonempty_questions_answered(questions: list, answers: list[str]) -> bool:
    """True when every non-empty question row has a successful answer."""
    for i, q in enumerate(questions):
        qtext = str(q).strip() if q is not None else ""
        if not qtext:
            continue
        if i >= len(answers):
            return False
        if not _is_successful_answer(answers[i]):
            return False
    return True


def _derive_status_after_answers(questions: list, answers: list[str], previous: str | None) -> str:
    """
    Set Completed when all substantive questions have successful answers.
    If no longer complete, downgrade from Completed to Draft; otherwise keep prior workflow status.
    """
    prev = (previous or "").strip() or "Draft"
    if _all_nonempty_questions_answered(questions, answers):
        return "Completed"
    if prev == "Completed":
        return "Draft"
    return prev


def _display_recipients(db: DbSession, row: RFPQuestion, owner: User | None) -> list[str]:
    """
    Display labels for avatars: owner, collaborators (by user id), optional legacy recipients JSON.
    """
    seen_lower: set[str] = set()
    out: list[str] = []

    def add_label(s: str) -> None:
        s = str(s).strip()
        if not s:
            return
        key = s.lower()
        if key in seen_lower:
            return
        seen_lower.add(key)
        out.append(s)

    if owner is not None:
        add_label((owner.name or "").strip() or (owner.email or "").strip() or "User")
    for uid in _parse_collab_ids(getattr(row, "collaborator_user_ids", None)):
        u = db.get(User, uid)
        if u is not None:
            add_label((u.name or "").strip() or (u.email or "").strip() or uid)
    try:
        legacy = json.loads(row.recipients) if row.recipients else []
    except (json.JSONDecodeError, TypeError):
        legacy = []
    if isinstance(legacy, list):
        for x in legacy:
            add_label(str(x).strip())
    if out:
        return out
    if owner is not None:
        return [(owner.name or "").strip() or (owner.email or "").strip() or "User"]
    return []


def _ensure_rfp_conversation_id(db: DbSession, row: RFPQuestion) -> str:
    """One conversation_id per Excel/RFP — used to group all search_queries from bulk answer generation."""
    cid = getattr(row, "conversation_id", None)
    if cid and str(cid).strip():
        return str(cid).strip()
    new_id = generate_conversation_id()
    row.conversation_id = new_id
    db.add(row)
    db.commit()
    db.refresh(row)
    return new_id


def _average_accuracy_ratio(confidence: list) -> float | None:
    """Mean of per-question confidence values in [0, 1], or None if nothing usable. Accepts 0–100 as well."""
    if not confidence:
        return None
    vals: list[float] = []
    for x in confidence:
        try:
            v = float(x)
        except (TypeError, ValueError):
            continue
        if 0.0 <= v <= 1.0:
            vals.append(v)
        elif 1.0 < v <= 100.0:
            vals.append(v / 100.0)
    if not vals:
        return None
    return sum(vals) / len(vals)


def _is_unanswered_for_metrics(a: str | None) -> bool:
    """Aligns with frontend isUnansweredAnswer for counts in emails."""
    s = (a or "").strip()
    if not s or s == NO_CONTEXT_MESSAGE:
        return True
    if s.startswith(UNANSWERED_PREFIX):
        return True
    if s.startswith("[Error:"):
        return True
    return False


def _count_answered_unanswered(questions: list, raw_answers: list[str]) -> tuple[int, int]:
    n = len(questions)
    answered = 0
    unanswered = 0
    for i in range(n):
        a = raw_answers[i] if i < len(raw_answers) else ""
        if _is_unanswered_for_metrics(a):
            unanswered += 1
        else:
            answered += 1
    return answered, unanswered


@router.get("", response_model=dict)
async def list_rfp_questions(
    db: DbSession,
    current_user: CurrentUser,
    skip: int = Query(0, ge=0, description="Number of records to skip"),
    limit: int = Query(20, ge=1, le=100, description="Max records per page"),
    user_id: str | None = Query(None, description="Filter by user ID (optional; admins only)"),
    status: str | None = Query(None, description="Filter by status (e.g. Draft, Sent)"),
):
    """
    List RFP questions with pagination.
    Returns items and total count.
    """
    q = select(RFPQuestion, User).join(User, RFPQuestion.user_id == User.id)
    count_q = select(func.count()).select_from(RFPQuestion)
    if current_user.role not in (UserRole.admin, UserRole.manager):
        q = q.where(_rfp_accessible_filter(current_user))
        count_q = count_q.where(_rfp_accessible_filter(current_user))
    elif user_id is not None:
        q = q.where(RFPQuestion.user_id == user_id)
        count_q = count_q.where(RFPQuestion.user_id == user_id)
    if status is not None and status.strip():
        q = q.where(RFPQuestion.status == status.strip())
        count_q = count_q.where(RFPQuestion.status == status.strip())
    total = db.execute(count_q).scalar_one()
    q = q.order_by(RFPQuestion.last_activity_at.desc()).offset(skip).limit(limit)
    rows = db.execute(q).all()
    items = []
    for r, owner in rows:
        items.append({
            "id": r.id,
            "rfpid": r.rfpid,
            "name": r.name,
            "user_id": r.user_id,
            "created_at": r.created_at.isoformat() if r.created_at else None,
            "last_activity_at": r.last_activity_at.isoformat() if r.last_activity_at else None,
            "recipients": _display_recipients(db, r, owner),
            "collaborator_user_ids": _parse_collab_ids(getattr(r, "collaborator_user_ids", None)),
            "conversation_id": getattr(r, "conversation_id", None),
            "status": r.status,
        })
    return {"items": items, "total": total}


@router.get("/{rfpid}", response_model=dict)
async def get_rfp(rfpid: str, db: DbSession, current_user: CurrentUser):
    """Get a single RFP by rfpid (full details including questions and answers)."""
    row = db.execute(select(RFPQuestion).where(RFPQuestion.rfpid == rfpid)).scalars().one_or_none()
    if not row:
        raise HTTPException(status_code=404, detail="RFP not found")
    _require_rfp_access(current_user, row)
    questions = json.loads(row.questions) if row.questions else []
    answers = json.loads(row.answers) if row.answers else []
    confidence = _confidence_as_array(getattr(row, "confidence", None))
    owner = db.execute(select(User).where(User.id == row.user_id)).scalars().one_or_none()
    recipients = _display_recipients(db, row, owner)
    conv_id = _ensure_rfp_conversation_id(db, row)
    # When no context was found, answer is empty; return user-facing message
    answers_for_response = _answers_for_response(answers)
    avg = _average_accuracy_ratio(confidence)
    return {
        "id": row.id,
        "rfpid": row.rfpid,
        "name": row.name,
        "user_id": row.user_id,
        "created_at": row.created_at.isoformat() if row.created_at else None,
        "last_activity_at": row.last_activity_at.isoformat() if row.last_activity_at else None,
        "recipients": recipients,
        "collaborator_user_ids": _parse_collab_ids(getattr(row, "collaborator_user_ids", None)),
        "conversation_id": conv_id,
        "status": row.status,
        "questions": questions,
        "answers": answers_for_response,
        "confidence": confidence,
        "average_accuracy": avg,
    }


@router.delete("/{rfpid}", response_model=dict)
async def delete_rfp(rfpid: str, db: DbSession, current_user: CurrentUser):
    """Delete an RFP by rfpid. Permanently removes the record."""
    row = db.execute(select(RFPQuestion).where(RFPQuestion.rfpid == rfpid)).scalars().one_or_none()
    if not row:
        raise HTTPException(status_code=404, detail="RFP not found")
    _require_rfp_owner_or_privileged(current_user, row)
    db.delete(row)
    db.commit()
    return {"message": "RFP deleted", "rfpid": rfpid}


class UpdateCollaboratorsBody(BaseModel):
    """Replace collaborator list (user ids). Owner cannot be listed; unknown ids are rejected."""
    collaborator_user_ids: list[str]


@router.patch("/{rfpid}/collaborators", response_model=dict)
async def update_rfp_collaborators(
    rfpid: str,
    body: UpdateCollaboratorsBody,
    db: DbSession,
    current_user: CurrentUser,
):
    """Set who may access this RFP besides the owner (My RFPs list + view/edit answers). Owner or admin only."""
    row = db.execute(select(RFPQuestion).where(RFPQuestion.rfpid == rfpid)).scalars().one_or_none()
    if not row:
        raise HTTPException(status_code=404, detail="RFP not found")
    _require_rfp_owner_or_privileged(current_user, row)
    owner_id = row.user_id
    cleaned: list[str] = []
    for uid in body.collaborator_user_ids or []:
        u = (uid or "").strip()
        if not u or u == owner_id:
            continue
        if db.get(User, u) is None:
            raise HTTPException(status_code=400, detail=f"Unknown user id: {u}")
        cleaned.append(u)
    row.collaborator_user_ids = _format_collab_ids(cleaned)
    row.last_activity_at = datetime.now(timezone.utc)
    db.add(row)
    db.commit()
    db.refresh(row)
    owner = db.execute(select(User).where(User.id == row.user_id)).scalars().one_or_none()
    return {
        "rfpid": row.rfpid,
        "collaborator_user_ids": _parse_collab_ids(row.collaborator_user_ids),
        "recipients": _display_recipients(db, row, owner),
    }


class UpdateAnswersBody(BaseModel):
    """Request body for updating answers array (one answer per question, same order). Optional confidence array (numbers, same order)."""
    answers: list[str]
    confidence: list[float] | None = None  # optional: one confidence value per question, same row order


@router.patch("/{rfpid}/answers", response_model=dict)
async def update_rfp_answers(
    rfpid: str,
    body: UpdateAnswersBody,
    db: DbSession,
    background_tasks: BackgroundTasks,
    current_user: CurrentUser,
):
    """
    Update the answers array for an RFP (by rfpid).
    answers must be a list of strings, in the same order as questions.
    """
    row = db.execute(select(RFPQuestion).where(RFPQuestion.rfpid == rfpid)).scalars().one_or_none()
    if not row:
        raise HTTPException(status_code=404, detail="RFP not found")
    _require_rfp_access(current_user, row)
    questions = json.loads(row.questions) if row.questions else []
    previous_status = (row.status or "").strip() or "Draft"
    answers_json = json.dumps(body.answers)
    row.answers = answers_json
    if body.confidence is not None:
        # Store only array of numbers
        row.confidence = json.dumps([float(x) for x in list(body.confidence)])
    row.status = _derive_status_after_answers(questions, body.answers, row.status)
    row.last_activity_at = datetime.now(timezone.utc)
    db.add(row)
    db.commit()
    db.refresh(row)
    confidence_out = _confidence_as_array(row.confidence)
    # When no context was found, answer is empty; return user-facing message
    answers_for_response = _answers_for_response(body.answers)
    avg = _average_accuracy_ratio(confidence_out)

    if (
        row.status == "Completed"
        and previous_status != "Completed"
    ):
        user = db.execute(select(User).where(User.id == row.user_id)).scalars().one_or_none()
        to_email = (user.email or "").strip() if user else ""
        if to_email:
            acc_pct = round(avg * 100) if avg is not None else None
            ans_n, unans_n = _count_answered_unanswered(questions, body.answers)
            excel_b: bytes = b""
            try:
                excel_b = build_qa_excel_bytes(questions, body.answers)
            except Exception:
                logger.exception("Failed to build Q&A Excel for completion email rfpid=%s", rfpid)
            background_tasks.add_task(
                send_rfp_completion_notification,
                to_email=to_email,
                user_name=(user.name or "").strip() or "there",
                rfp_title=(row.name or "").strip() or "Your RFP",
                rfpid=rfpid,
                accuracy_pct=acc_pct,
                answered=ans_n,
                unanswered=unans_n,
                excel_bytes=excel_b,
            )

    return {
        "rfpid": row.rfpid,
        "id": row.id,
        "answers": answers_for_response,
        "confidence": confidence_out,
        "status": row.status,
        "average_accuracy": avg,
        "last_activity_at": row.last_activity_at.isoformat() if row.last_activity_at else None,
    }


ALLOWED_EXTENSIONS = {".xlsx", ".xls", ".csv"}
ALLOWED_CONTENT_TYPES = {
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    "application/vnd.ms-excel",
    "text/csv",
    "application/csv",
}


def _extract_questions_from_csv(content: bytes) -> list[str]:
    """Extract column A (first column) from CSV content."""
    text = content.decode("utf-8-sig")  # utf-8-sig handles BOM
    reader = csv.reader(io.StringIO(text))
    questions: list[str] = []
    for row in reader:
        if row and row[0]:
            val = str(row[0]).strip()
            if val:
                questions.append(val)
    return questions


def _extract_questions_from_excel(content: bytes) -> list[str]:
    """Extract column A (first column) from Excel content."""
    wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    ws = wb.active
    questions: list[str] = []
    for row in ws.iter_rows(min_row=1, max_col=1):
        cell = row[0]
        if cell.value is not None:
            val = str(cell.value).strip()
            if val:
                questions.append(val)
    return questions


def _extract_questions(file: UploadFile, body: bytes) -> list[str]:
    """Extract questions from Excel or CSV file (column A)."""
    filename = (file.filename or "").lower()
    content_type = (file.content_type or "").lower()

    if filename.endswith(".csv") or "csv" in content_type:
        return _extract_questions_from_csv(body)
    if filename.endswith((".xlsx", ".xls")) or "spreadsheet" in content_type or "excel" in content_type:
        return _extract_questions_from_excel(body)

    # Fallback by extension
    if any(filename.endswith(ext) for ext in ALLOWED_EXTENSIONS):
        if ".csv" in filename:
            return _extract_questions_from_csv(body)
        return _extract_questions_from_excel(body)

    raise HTTPException(
        status_code=400,
        detail="Unsupported file format. Use Excel (.xlsx, .xls) or CSV.",
    )


@router.post("/import", response_model=dict)
async def import_questions(
    db: DbSession,
    current_user: CurrentUser,
    user_id: str = Form(..., description="User ID (UUID) who is importing"),
    file: UploadFile = File(..., description="Excel or CSV file with questions in column A"),
):
    """
    Import questions from Excel or CSV.
    Extracts column A as list of questions, generates rfpid, and stores in rfpquestions table.
    """
    logger.info("RFP questions import: filename=%s user_id=%s", file.filename, user_id)

    if current_user.role not in (UserRole.admin, UserRole.manager) and user_id != current_user.id:
        raise HTTPException(status_code=403, detail="You can only import RFPs for your own account")

    # Validate user exists
    user = db.execute(select(User).where(User.id == user_id)).scalars().one_or_none()
    if not user:
        raise HTTPException(status_code=404, detail="User not found")

    body = await file.read()
    if not body:
        raise HTTPException(status_code=400, detail="File is empty")

    questions = _extract_questions(file, body)
    if not questions:
        raise HTTPException(status_code=400, detail="No questions found in column A")

    rfpid = generate_rfpid()
    now = datetime.now(timezone.utc)
    name = (file.filename or "Untitled RFP").rsplit(".", 1)[0]  # strip extension
    if not name.strip():
        name = "Untitled RFP"
    questions_json = json.dumps(questions)
    answers_json = json.dumps([])
    confidence_json = json.dumps([])  # one number per question, same order; empty until populated
    recipients_json = json.dumps([])
    conv_id = generate_conversation_id()

    record = RFPQuestion(
        rfpid=rfpid,
        user_id=user_id,
        name=name[:512],
        created_at=now,
        last_activity_at=now,
        conversation_id=conv_id,
        recipients=recipients_json,
        status="Draft",
        questions=questions_json,
        answers=answers_json,
        confidence=confidence_json,
    )
    db.add(record)
    db.commit()
    db.refresh(record)

    return {
        "rfpid": rfpid,
        "id": record.id,
        "name": record.name,
        "question_count": len(questions),
        "last_activity_at": record.last_activity_at.isoformat() if record.last_activity_at else None,
        "recipients": _display_recipients(db, record, user),
        "conversation_id": conv_id,
        "status": record.status,
    }
